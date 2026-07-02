import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

# Files to format
matrix_file = r"c:\Users\anura\Downloads\LetzRyd_Forms_Registry_Matrix.xlsx"
flat_file = r"c:\Users\anura\Downloads\LetzRyd_Forms_Registry.xlsx"

def format_matrix():
    print("Formatting LetzRyd_Forms_Registry_Matrix.xlsx...")
    
    # 1. Read existing data to transpose
    df_raw = pd.read_excel(matrix_file)
    
    # Columns in raw matrix:
    # form name | Walk-In Form | Onboarding Form | Adjustment Form | Allocation Form | Expenses Form
    # Row 0: business use case
    # Row 1: Variables (headers for Field Label | Input Type | Req. etc. - we ignore this placeholder row)
    # Row 2 onwards: field labels (e.g. Visitor Classification *, Driver Name *, etc.)
    
    forms = ["Walk-In Form", "Onboarding Form", "Adjustment Form", "Allocation Form", "Expenses Form"]
    use_cases = {
        "Walk-In Form": df_raw.loc[0, "Walk-In Form"],
        "Onboarding Form": df_raw.loc[0, "Onboarding Form"],
        "Adjustment Form": df_raw.loc[0, "Adjustment Form"],
        "Allocation Form": df_raw.loc[0, "Allocation Form"],
        "Expenses Form": df_raw.loc[0, "Expenses Form"]
    }
    
    # Extract fields vertically
    form_fields = {}
    for form in forms:
        # Get values from row 2 onwards (which correspond to index 1 in dataframe, excluding the Variables header at index 0)
        # Note: df_raw index 0 is row 2 of spreadsheet ('business use case')
        # df_raw index 1 is row 3 of spreadsheet ('Variables' row - we skip this)
        # df_raw index 2 onwards are the fields
        fields = df_raw.loc[2:, form].dropna().tolist()
        form_fields[form] = fields
        
    # Determine max fields (to make space for at least 30 fields)
    max_fields = 30
    
    # Create the transposed dataframe structure
    columns = ["Form Name", "Business Use Case"] + [f"Field {i}" for i in range(1, max_fields + 1)]
    
    # Build rows
    rows_data = []
    for form in forms:
        row = [form, use_cases[form]]
        fields = form_fields[form]
        # Pad with None up to max_fields
        padded_fields = fields + [None] * (max_fields - len(fields))
        row.extend(padded_fields)
        rows_data.append(row)
        
    # 2. Write to new openpyxl workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portal Forms Matrix"
    ws.views.sheetView[0].showGridLines = True
    
    # Styling variables
    font_family = "Inter"
    navy_fill = PatternFill(start_color="0A1650", end_color="0A1650", fill_type="solid")
    teal_fill = PatternFill(start_color="1AB394", end_color="1AB394", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Row 1: Merged Main Header
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "LetzRyd Portal Forms Variable Registry"
    title_cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Color the merged cells in Row 1
    for col in range(2, len(columns) + 1):
        ws.cell(row=1, column=col).fill = navy_fill
        
    # Row 2: Columns Header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = col_name
        cell.font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        cell.fill = teal_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28
    
    # Row 3 onwards: Transposed Data rows
    for row_idx, rdata in enumerate(rows_data, 3):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(rdata, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = Font(name=font_family, size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
            # Format "Form Name" and "Business Use Case" slightly bold
            if col_idx in [1, 2]:
                cell.font = Font(name=font_family, size=10, bold=True)
                
    # Add empty rows up to row 35 (to allow adding 20-30 forms manually in future)
    for row_idx in range(len(rows_data) + 3, 40):
        ws.row_dimensions[row_idx].height = 24
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name=font_family, size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col_idx in [1, 2]:
                cell.font = Font(name=font_family, size=10, bold=True)

    # Freeze the first two rows
    ws.freeze_panes = "A3"
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't size based on merged header row 1
        for cell in col[1:]:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Explicitly make Form Name & Business Use Case wider
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 45
    
    wb.save(matrix_file)
    print("Matrix file saved successfully.")

def format_flat():
    print("Formatting LetzRyd_Forms_Registry.xlsx...")
    # 1. Read existing data
    df = pd.read_excel(flat_file)
    
    # 2. Capitalize the first letter of variable details if required
    # Wait, "caps the first word for variable details"
    # Column 'Variable/Field Name' (e.g. event_date -> Event_date or Event Date?)
    # Let's capitalize the first letter of 'Variable/Field Name' or 'Display Label' columns.
    # Display labels: e.g. "visitor classification" -> "Visitor classification" or "Visitor Classification"
    # Let's do both to be safe: capitalize the Display Label first word if it exists.
    if "Variable/Field Name" in df.columns:
        df["Variable/Field Name"] = df["Variable/Field Name"].apply(
            lambda x: str(x).capitalize() if isinstance(x, str) else x
        )
    if "Display Label" in df.columns:
        df["Display Label"] = df["Display Label"].apply(
            lambda x: str(x).capitalize() if isinstance(x, str) else x
        )
        
    # Write to new openpyxl workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Inter"
    navy_fill = PatternFill(start_color="0A1650", end_color="0A1650", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Headers
    headers = df.columns.tolist()
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    
    # Data Rows
    for row_idx, row_data in enumerate(df.values, 2):
        ws.row_dimensions[row_idx].height = 22
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = Font(name=font_family, size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
        
    wb.save(flat_file)
    print("Flat list file saved successfully.")

if __name__ == "__main__":
    format_matrix()
    format_flat()
    print("All formatting operations completed.")
